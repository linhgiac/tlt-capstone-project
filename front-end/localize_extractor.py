import pandas as pd
import openpyxl
import os
import numpy as np
import json


def write_json(data, filename):
    with open(filename, "w", encoding='utf-8') as f:
        json.dump(data, f, indent=4, ensure_ascii=False)


def main():

    sheetnames = openpyxl.load_workbook('TLTLocalize.xlsx').sheetnames

    support_language = ['en', 'de', 'fr', 'ja', 'vi']
    for sheetname in sheetnames:
        df = pd.read_excel('TLTLocalize.xlsx', sheet_name=sheetname)
        all_language = df.head(0)
        all_key = df['key']
        for language in all_language:
            if(language in support_language):
                language_folder = "public/locales/" + language
                if not os.path.exists(language_folder):
                    os.makedirs(language_folder)
                all_key_value = {}
                all_value = df[language]
                for key, value in zip(all_key, all_value):
                    all_key_value[key] = value
                write_json(all_key_value, language_folder + "/" + sheetname + ".json")


if __name__ == '__main__':
    main()
